import os
import glob
import logging
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Constants (source: docs/BUSINESS_RULES.md)
FUEL_EFFICIENCY_KM_PER_LITER = 17.23
FUEL_PRICE_PER_LITER = 23.24

# Traffic light thresholds
THRESHOLD_GREEN = 3000
THRESHOLD_YELLOW = 6000

# Raw column names
COL_DISTANCE  = "Distancia recorrida total,\nkm"
COL_AVG_SPEED = "Velocidad media,\nkm/h"
COL_MAX_SPEED = "Max. velocidad,\nkm/h"
COL_START     = "Inicio de movimiento"
COL_END       = "Final de movimiento"

REQUIRED_COLUMNS = [COL_DISTANCE, COL_AVG_SPEED, COL_MAX_SPEED, COL_START, COL_END]

# City name -> short code
CITY_CODES: dict[str, str] = {
    "AGUASCALIENTES": "ags",
    "CELAYA":         "cel",
    "QUERETARO":      "qro",
    "SAN JUAN":       "snjuan",
    "SILAO":          "silao",
    "S.L.P.":         "slp",
    "TOLUCA":         "tol",
    "ZACATECAS":      "zac",
}

DATA_FOLDERS: list[str] = os.getenv(
    "GPS_DATA_FOLDERS", "data_gps/dic_2021,data_gps/ene_2022"
).split(",")

# Spanish month abbreviations -> month number
MONTHS_ES: dict[str, str] = {
    "ene": "01", "feb": "02", "mar": "03", "abr": "04",
    "may": "05", "jun": "06", "jul": "07", "ago": "08",
    "sep": "09", "oct": "10", "nov": "11", "dic": "12",
}


def extract_city(file_path: str) -> str:
    """Extract city name from filename. Handles 'VOLVO CITY-MES' and 'CITY-MES' formats."""
    name = os.path.basename(file_path).replace("VOLVO ", "")
    for suffix in ["-DIC 2021", "  DIC 2021", "-ENERO 2022", ".ENERO 2022", "- ENERO 2022"]:
        name = name.replace(suffix, "")
    return name.replace(".xlsx", "").strip()


def find_files() -> list[str]:
    """Find all raw Excel files in DATA_FOLDERS."""
    files = []
    for folder in DATA_FOLDERS:
        files += glob.glob(os.path.join(folder.strip(), "*.xlsx"))
    return files


def _extract_time_string(col: pd.Series) -> pd.Series:
    """Extract 'HH:MM' from strings like '10:29 - Calle Villas...'"""
    return col.str.split(" - ").str[0]


def process_file(file_path: str, agency: str) -> pd.DataFrame:
    """Read Excel, clean data, return DataFrame with all trips.

    Each row = one trip. Columns include date, day, month, agency, unit,
    start_hour, end_hour, off_hours, distance, speed metrics.
    """
    try:
        wb = load_workbook(file_path, read_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open Excel file '{file_path}': {e}") from e

    sheets = wb.sheetnames
    wb.close()

    frames = []
    for sheet in sheets:
        if sheet == sheets[0] or sheet.endswith(" - 2"):
            continue

        try:
            df = pd.read_excel(file_path, sheet_name=sheet, header=4, usecols="A:G")
        except Exception as e:
            logger.warning("Skipping sheet '%s' in '%s': %s", sheet, file_path, e)
            continue

        # Validate required columns exist
        missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
        if missing:
            logger.warning(
                "Sheet '%s' in '%s' missing columns: %s — skipping",
                sheet, file_path, missing,
            )
            continue

        # --- Extract date from grouped header rows ---
        date_mask  = df[COL_START].astype(str).str.match(r"^\d{2}-\w+-\d{4}")
        raw_dates  = df.loc[date_mask, COL_START].str.extract(r"^(\d{2})-(\w+)-(\d{4})")
        normalized = raw_dates[0] + "-" + raw_dates[1].str.lower().map(MONTHS_ES) + "-" + raw_dates[2]
        df["date"] = pd.to_datetime(normalized, format="%d-%m-%Y", errors="coerce")
        df["date"] = df["date"].ffill()

        # --- Filter actual trip rows ---
        df = df.dropna(subset=[COL_START, COL_END, COL_DISTANCE])
        df = df[~df[COL_START].astype(str).str.contains("En total")]

        if df.empty:
            logger.warning("Sheet '%s' in '%s' has no valid trips — skipping", sheet, file_path)
            continue

        # --- Derive temporal columns from date ---
        df["day"]   = df["date"].dt.day_name()
        df["month"] = df["date"].dt.month_name()
        df["year"]  = df["date"].dt.year

        # --- Extract time strings ---
        df["start_time"] = _extract_time_string(df[COL_START])
        df["end_time"]   = _extract_time_string(df[COL_END])

        # --- Off-hours flag (after 19:00 or before 05:00) ---
        start_h = pd.to_datetime(df["start_time"], format="mixed", errors="coerce").dt.hour
        end_h   = pd.to_datetime(df["end_time"],   format="mixed", errors="coerce").dt.hour
        df["off_hours"] = ((start_h >= 19) | (end_h <= 5)).map({True: "yes", False: "no"})

        # --- Convert timedelta to minutes ---
        df["travel_time_min"] = df["Tiempo de viaje"] / pd.Timedelta(minutes=1)
        df["idle_time_min"] = df["Tiempo de inactividad"] / pd.Timedelta(minutes=1)

        # --- Identity columns ---
        df["unit"]   = sheet
        df["agency"] = agency

        frames.append(df)

    if not frames:
        raise ValueError(f"No valid trip sheets found in '{file_path}'")

    return pd.concat(frames, ignore_index=True)
