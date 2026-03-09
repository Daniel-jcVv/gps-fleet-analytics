import logging
import pandas as pd
from openpyxl.styles import Font, Alignment, numbers
from .extract import COL_DISTANCE, COL_AVG_SPEED, COL_MAX_SPEED

logger = logging.getLogger(__name__)

# Columns to include in the "data" sheet (trip-level, matches MASTER original)
DATA_COLUMNS = [
    "agency", "unit", "date", "day", "month", "year",
    "start_time", "end_time",
    COL_DISTANCE, COL_AVG_SPEED, COL_MAX_SPEED,
    "off_hours", "travel_time_min", "idle_time_min"
]

# Column display widths (characters)
COLUMN_WIDTHS: dict[str, int] = {
    "agency": 18,
    "unit": 12,
    "date": 14,
    "day": 12,
    "month": 12,
    "year": 8,
    "start_time": 12,
    "end_time": 12,
    "distance_km": 14,
    "avg_speed_kmh": 16,
    "max_speed_kmh": 16,
    "off_hours": 12,
    "travel_time_min": 16,
    "idle_time_min": 14,
}

# Number format per column
NUMBER_FORMATS: dict[str, str] = {
    "distance_km": "#,##0.00",
    "avg_speed_kmh": "#,##0.00",
    "max_speed_kmh": "#,##0.00",
    "travel_time_min": "#,##0.0",
    "idle_time_min": "#,##0.0",
}


def _apply_formatting(ws) -> None:
    """Apply header style, column widths, and number formats to worksheet."""
    # Header: bold, centered
    header_font = Font(bold=True)
    center = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center

    # Column widths and number formats
    for col_idx, cell in enumerate(ws[1], start=1):
        col_name = cell.value
        col_letter = cell.column_letter

        # Width
        width = COLUMN_WIDTHS.get(col_name, 14)
        ws.column_dimensions[col_letter].width = width

        # Number format for data rows
        fmt = NUMBER_FORMATS.get(col_name)
        if fmt:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = fmt


def export_report(df_trips: pd.DataFrame, output_path: str) -> str:
    """Export trip-level data to Excel with formatted 'data' sheet."""
    cols = [c for c in DATA_COLUMNS if c in df_trips.columns]
    df_data = df_trips[cols].copy()
    df_data = df_data.rename(columns={
        "Distancia recorrida total,\nkm": "distance_km",
        "Velocidad media,\nkm/h":        "avg_speed_kmh",
        "Max. velocidad,\nkm/h":         "max_speed_kmh",
    })
    df_data.columns = [c.replace("\n", " ") for c in df_data.columns]

    # Format date as DD/MM/YYYY
    df_data["date"] = df_data["date"].apply(
        lambda d: d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else d
    )

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_data.to_excel(writer, sheet_name="data", index=False)
            _apply_formatting(writer.sheets["data"])
    except PermissionError:
        raise PermissionError(
            f"Cannot write to '{output_path}' — file may be open in another program"
        )
    except Exception as e:
        raise IOError(f"Failed to write Excel report '{output_path}': {e}") from e

    logger.info("Exported %d trips to '%s'", len(df_data), output_path)
    return output_path
