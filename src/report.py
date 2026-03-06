from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from .ingestion import COL_DISTANCE, COL_AVG_SPEED, COL_MAX_SPEED

FILL_COLORS = {
    "green":  PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "yellow": PatternFill(start_color="FFEB96", end_color="FFEB9C", fill_type="solid"),
    "red":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

# Columns to include in the "data" sheet (trip-level, matches MASTER original)
DATA_COLUMNS = [
    "agency", "unit", "date", "day", "month", "year",
    "start_time", "end_time",
    COL_DISTANCE, COL_AVG_SPEED, COL_MAX_SPEED,
    "off_hours",
]


def export_report(df_trips, output_path):
    """Export trip-level data to Excel with single 'data' sheet."""
    # --- Sheet: data ---
    cols = [c for c in DATA_COLUMNS if c in df_trips.columns]
    df_data = df_trips[cols].copy()
    df_data = df_data.rename(columns={
        "Distancia recorrida total,\nkm": "distance_km",
        "Velocidad media,\nkm/h":        "avg_speed_kmh",
        "Max. velocidad,\nkm/h":         "max_speed_kmh",
    })
    df_data.columns = [c.replace("\n", " ") for c in df_data.columns]

    # Format date as DD/MM/YYYY
    df_data["date"] = df_data["date"].apply(
        lambda d: d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else d
    )

    with __import__("pandas").ExcelWriter(output_path, engine="openpyxl") as writer:
        df_data.to_excel(writer, sheet_name="data", index=False)

    return output_path
