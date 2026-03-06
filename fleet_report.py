import os
import warnings
import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Constants (source: BUSINESS_RULES.md)
FUEL_EFFICIENCY_KM_PER_LITER = 17.23
FUEL_PRICE_PER_LITER = 21

# Raw column names
COL_DISTANCE = "Distancia recorrida total,\nkm"
COL_AVG_SPEED = "Velocidad media,\nkm/h"
COL_MAX_SPEED = "Max. velocidad,\nkm/h"
COL_START = "Inicio de movimiento"
COL_END = "Final de movimiento"


def extract_time_string(col):
    """Extract 'HH:MM' from strings like '10:29 - Calle Villas...'"""
    return col.str.split(" - ").str[0]


def process_file(file_path):
    """Read Excel, clean data, return DataFrame with all trips."""
    wb = load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    frames = []
    for sheet in sheets:
        if sheet == sheets[0] or sheet.endswith(" - 2"):
            continue

        df = pd.read_excel(file_path, sheet_name=sheet, header=4, usecols="A:G")
        df = df.dropna(subset=[COL_START, COL_END, COL_DISTANCE])
        df = df[~df[COL_START].astype(str).str.contains("En total")]
        df["unit"] = sheet
        frames.append(df)

    df_total = pd.concat(frames, ignore_index=True)

    df_total["start_hour"] = pd.to_datetime(extract_time_string(df_total[COL_START]), format="mixed").dt.hour
    df_total["end_hour"]   = pd.to_datetime(extract_time_string(df_total[COL_END]),   format="mixed").dt.hour
    df_total["off_hours"]  = (df_total["start_hour"] >= 19) | (df_total["end_hour"] <= 5)

    return df_total


def build_summary(df):
    """Group by unit, calculate KPIs, fuel cost, traffic light, night metrics."""
    summary = df.groupby("unit").agg(
        distance_km=(COL_DISTANCE, "sum"),
        avg_speed_kmh=(COL_AVG_SPEED, "mean"),
        max_speed_kmh=(COL_MAX_SPEED, "max"),
        off_hours_trips=("off_hours", "sum"),
    )

    summary["liters"]  = summary["distance_km"] / FUEL_EFFICIENCY_KM_PER_LITER
    summary["cost"]    = summary["liters"] * FUEL_PRICE_PER_LITER

    # Night metrics
    df_night = df[df["off_hours"]]
    night = df_night.groupby("unit")[COL_DISTANCE].sum().rename("distance_night")
    summary = summary.join(night)
    summary["cost_night"] = (summary["distance_night"] / FUEL_EFFICIENCY_KM_PER_LITER) * FUEL_PRICE_PER_LITER

    # Traffic light
    def traffic_light(km):
        if km < 3000:   return "green"
        elif km <= 6000: return "yellow"
        else:            return "red"

    summary["status"] = summary["distance_km"].apply(traffic_light)

    return summary.sort_values("cost", ascending=False)


def export_report(df, output_path):
    """Export summary to Excel with traffic light row formatting."""
    df.to_excel(output_path, sheet_name="summary")

    wb = load_workbook(output_path)
    ws = wb["summary"]

    fill_colors = {
        "green":  PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "yellow": PatternFill(start_color="FFEB96", end_color="FFEB9C", fill_type="solid"),
        "red":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    status_col = df.columns.get_loc("status") + 2  # +2: 1-indexed + index col
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=status_col).value
        if status in fill_colors:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill_colors[status]

    wb.save(output_path)
    return output_path


CITY_CODES = {
    "AGUASCALIENTES": "ags",
    "CELAYA":         "cel",
    "QUERETARO":      "qro",
    "SAN JUAN":       "snjuan",
    "SILAO":          "silao",
    "S.L.P.":         "slp",
    "TOLUCA":         "tol",
    "ZACATECAS":      "zac",
}

DATA_FOLDERS = [
    "data_gps/dic_2021",
    "data_gps/ene_2022",
]


def extract_city(file_path):
    """Extract city name from filename. Handles 'VOLVO CITY-MES' and 'CITY-MES' formats."""
    name = os.path.basename(file_path).replace("VOLVO ", "")
    for suffix in ["-DIC 2021", "  DIC 2021", "-ENERO 2022", ".ENERO 2022", "- ENERO 2022"]:
        name = name.replace(suffix, "")
    return name.replace(".xlsx", "").strip()


def main():
    os.makedirs("data_gps/output", exist_ok=True)

    files = []
    for folder in DATA_FOLDERS:
        files += glob.glob(os.path.join(folder, "*.xlsx"))

    if not files:
        print("Error: No .xlsx files found in data folders")
        return

    processed = 0
    failed = 0

    for f in files:
        city = extract_city(f)
        code = CITY_CODES.get(city)

        if not code:
            print(f"  SKIPPED: Unknown city '{city}' in {os.path.basename(f)}")
            failed += 1
            continue

        try:
            period = os.path.basename(os.path.dirname(f))
            df = process_file(f)
            summary = build_summary(df)
            filename = f"fleet_report_{code}_{period}.xlsx"
            output_path = os.path.join("data_gps/output", filename)
            export_report(summary, output_path)
            print(f"  OK: {filename} ({len(df)} trips)")
            processed += 1
        except Exception as e:
            print(f"  FAILED: {os.path.basename(f)} - {e}")
            failed += 1

    print(f"\nDone. {processed} reports saved, {failed} failed.")


if __name__ == "__main__":
    main()
